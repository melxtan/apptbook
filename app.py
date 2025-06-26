import streamlit as st
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime, timedelta, time

st.set_page_config(layout="wide")
st.title("Automated REDCap & Excel Workflow")

excel_file = st.file_uploader("Upload Excel File (.xlsx)", type=["xlsx"])
api_keys = [st.secrets["redcap_api_1"], st.secrets["redcap_api_2"]]

def parse_json_to_excel(json_data, ws_mrn):
    existing_case_ids = set()
    for row in ws_mrn.iter_rows(min_row=2, max_row=ws_mrn.max_row, min_col=9, max_col=9, values_only=True):
        if row[0]:
            existing_case_ids.add(str(row[0]).strip())
    new_mrn_count = 0
    i = ws_mrn.max_row + 1
    for record in json_data:
        incoming_case_id = str(record.get("full_case_id", "")).strip()
        if record.get("mrn") and incoming_case_id not in existing_case_ids:
            fields = [
                "mrn", "case_id", "redcap_event_name", "redcap_repeat_instrument",
                "redcap_repeat_instance", "country_origin", "first_responder",
                "internal_referral", "full_case_id", "arm_label", "email",
                "num_appt", "payer_type", "other_refer", "pt_fn", "pt_ln",
                "pt_dob", "today", "service_line"
            ]
            for col_idx, field in enumerate(fields, 1):
                ws_mrn.cell(row=i, column=col_idx, value=record.get(field, ""))
            new_mrn_count += 1
            i += 1
    return new_mrn_count

def sheet_to_df(ws):
    data = list(ws.values)
    if not data or not data[0]:
        return pd.DataFrame()
    cols = [str(c) if c is not None else f'col{ix+1}' for ix, c in enumerate(data[0])]
    return pd.DataFrame(data[1:], columns=cols)

def df_to_sheet(ws, df, highlight_rows=None):
    # Overwrite worksheet with DataFrame content, with optional highlighting.
    ws.delete_rows(2, ws.max_row-1)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if highlight_rows and (r_idx-2) in highlight_rows:
                cell.font = Font(color='FF0000')

def move_routine_to_newop(df_newop, df_routine, df_mrn):
    # 1. Extract routine rows to move (rows 6+, A-V)
    df_routine_to_move = df_routine.iloc[5:, :22].copy()
    df_routine_to_move = df_routine_to_move.dropna(how='all')
    if df_routine_to_move.empty:
        return df_newop, df_routine, df_mrn, []
    
    # 2. Expand routine to match newop columns
    missing_cols = [col for col in df_newop.columns if col not in df_routine_to_move.columns]
    for col in missing_cols:
        df_routine_to_move[col] = ""
    df_routine_to_move = df_routine_to_move[df_newop.columns]

    # 3. Concatenate
    before_rows = len(df_newop)
    df_newop = pd.concat([df_newop, df_routine_to_move], ignore_index=True)
    # Mark moved rows (for highlighting in Excel)
    highlight_rows = list(range(before_rows, before_rows + len(df_routine_to_move)))

    # 4. VLOOKUP: Fill X (23), Z (25), AA (26)
    for idx in highlight_rows:
        lookup_val = df_newop.iloc[idx, 1]  # Column B: MRN or ID
        # MRN tab: column 0: MRN, 1: case_id, ..., 5: col F, 6: col G
        vlookup_row = df_mrn[df_mrn.iloc[:, 0] == lookup_val]
        if not vlookup_row.empty:
            if df_newop.shape[1] > 23:
                df_newop.iat[idx, 23] = vlookup_row.iloc[0, 1] if df_mrn.shape[1] > 1 else ""
            if df_newop.shape[1] > 25:
                df_newop.iat[idx, 25] = vlookup_row.iloc[0, 5] if df_mrn.shape[1] > 5 else ""
            if df_newop.shape[1] > 26:
                df_newop.iat[idx, 26] = vlookup_row.iloc[0, 6] if df_mrn.shape[1] > 6 else ""
    # 5. Remove duplicates (A:AA)
    df_newop = df_newop.drop_duplicates(subset=list(df_newop.columns[:27]), keep='first').reset_index(drop=True)

    # 6. Clear Routine rows 6+ (A–V)
    df_routine.iloc[5:, :22] = ""

    return df_newop, df_routine, df_mrn, highlight_rows

if excel_file:
    excel_bytes = BytesIO(excel_file.read())
    wb = load_workbook(excel_bytes)
    ws_newop = wb["New OP"]
    ws_routine = wb["Routine"]
    ws_mrn = wb["MRN"]

    df_newop = sheet_to_df(ws_newop)
    df_routine = sheet_to_df(ws_routine)
    df_mrn = sheet_to_df(ws_mrn)

    if st.button("Refresh MRN sheet from REDCap (API)"):
        total_new = 0
        for key in api_keys:
            response = requests.post(
                "https://redcap.med.usc.edu/api/",
                data={
                    "token": key,
                    "content": "record",
                    "format": "json",
                    "type": "flat",
                    "rawOrLabel": "label"
                }
            )
            if response.ok:
                data = response.json()
                new_mrn = parse_json_to_excel(data, ws_mrn)
                total_new += new_mrn
        st.success(f"MRN sheet refreshed. {total_new} new MRNs imported.")

    if st.button("Run Outpatient Routine Process"):
        df_newop_updated, df_routine_updated, df_mrn_updated, highlight_rows = move_routine_to_newop(df_newop, df_routine, df_mrn)
        df_to_sheet(ws_newop, df_newop_updated, highlight_rows)
        df_to_sheet(ws_routine, df_routine_updated)
        df_to_sheet(ws_mrn, df_mrn_updated)
        st.success(f"Routine → New OP processing done. {len(highlight_rows)} rows moved and highlighted.")

    output = BytesIO()
    wb.save(output)
    st.download_button(
        label="Download Processed Excel",
        data=output.getvalue(),
        file_name="processed_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.info("You can now upload this Excel to SharePoint if needed.")

else:
    st.info("Please upload your Excel file (with CSV data already pasted into Routine tab, A6).")
